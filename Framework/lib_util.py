import configparser

import openpyxl as xl

config = configparser.RawConfigParser()
configFilePath = ".\\Config\\lib_constants.ini"
config.read(configFilePath)


class ReadConfig:
    @staticmethod
    def getValue(constType, valueToGet):
        value_config = config.get(constType, valueToGet)
        return value_config


class ReadExcel:
    def loadDatafromExcel(filepath, tcName, dataset):
        workbook = xl.load_workbook(filepath)
        worksheet = workbook[tcName]

        dictionary = {}

        for row in range(1, worksheet.max_row + 1):
            key = worksheet.cell(row, 1).value
            value = worksheet.cell(row, 2).value
            dictionary[key] = value
        # print(dictionary)
        return dictionary
